import sys
import json
import time
import random
from bs4 import BeautifulSoup
import requests
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from urllib.parse import urljoin
import re
from openpyxl import Workbook
from datetime import datetime

class EmployeeSiteAnalyzer:
    def __init__(self, target_url):
        self.target_url = self.normalize_url(target_url)
        self.domain = self.extract_domain(target_url)
        self.employees = []
        self.driver = self.init_selenium()
        self.results = {
            'company_info': {},
            'employees': [],
            'related_sites': [],
            'social_media': {},
            'technologies': []
        }

    def normalize_url(self, url):
        if not url.startswith(('http://', 'https://')):
            return 'https://' + url
        return url

    def extract_domain(self, url):
        from urllib.parse import urlparse
        parsed = urlparse(url)
        return parsed.netloc.replace('www.', '')

    def init_selenium(self):
        try:
            chrome_options = Options()
            chrome_options.add_argument("--headless")
            chrome_options.add_argument("--disable-gpu")
            chrome_options.add_argument("--no-sandbox")
            return webdriver.Chrome(options=chrome_options)
        except Exception as e:
            print(f"Warning: Could not initialize Selenium - {e}")
            return None

    def get_company_info(self):
        print("[*] جمع معلومات الشركة...")
        try:
            # البحث في صفحة "من نحن" أو "عن الشركة"
            about_pages = ['about-us', 'about', 'company', 'من-نحن', 'عن-الشركة']
            
            for page in about_pages:
                try:
                    url = f"{self.target_url}/{page}"
                    response = requests.get(url, timeout=5)
                    if response.status_code == 200:
                        soup = BeautifulSoup(response.text, 'html.parser')
                        
                        # استخراج معلومات الشركة
                        company_name = soup.find('h1').text if soup.find('h1') else None
                        company_desc = soup.find('meta', attrs={'name': 'description'})['content'] if soup.find('meta', attrs={'name': 'description'}) else None
                        
                        self.results['company_info'] = {
                            'name': company_name,
                            'description': company_desc,
                            'about_page': url
                        }
                        break
                except:
                    continue
                    
            # إذا لم يتم العثور على صفحة "من نحن"، استخدم الصفحة الرئيسية
            if not self.results['company_info']:
                response = requests.get(self.target_url, timeout=5)
                soup = BeautifulSoup(response.text, 'html.parser')
                
                company_name = soup.title.text if soup.title else None
                company_desc = soup.find('meta', attrs={'name': 'description'})['content'] if soup.find('meta', attrs={'name': 'description'}) else None
                
                self.results['company_info'] = {
                    'name': company_name,
                    'description': company_desc,
                    'about_page': self.target_url
                }
                
        except Exception as e:
            print(f"Error getting company info: {e}")

    def find_team_page(self):
        print("[*] البحث عن صفحة الفريق أو الموظفين...")
        team_pages = ['team', 'staff', 'employees', 'leadership', 'الفريق', 'الموظفين', 'المدراء']
        
        for page in team_pages:
            try:
                url = f"{self.target_url}/{page}"
                response = requests.get(url, timeout=5)
                if response.status_code == 200:
                    soup = BeautifulSoup(response.text, 'html.parser')
                    
                    # تحليل صفحة الفريق
                    self.parse_team_page(soup, url)
                    return True
            except:
                continue
                
        return False

    def parse_team_page(self, soup, page_url):
        print("[*] تحليل صفحة الفريق...")
        # البحث عن عناصر الموظفين الشائعة
        employee_cards = soup.find_all(class_=re.compile('team-member|employee-card|staff-item'))
        
        if not employee_cards:
            employee_cards = soup.find_all('div', class_=re.compile('person|member'))
            
        if not employee_cards:
            employee_cards = soup.select('section.team li, div.team div')
            
        for card in employee_cards:
            try:
                employee = {}
                
                # استخراج اسم الموظف
                name = card.find(class_=re.compile('name|title'))
                if not name:
                    name = card.find('h3') or card.find('h2') or card.find('h4')
                employee['name'] = name.get_text().strip() if name else None
                
                # استخراج المنصب
                position = card.find(class_=re.compile('position|role|job-title'))
                if not position:
                    position = card.find('p') or card.find('span')
                employee['position'] = position.get_text().strip() if position else None
                
                # استخراج الصورة
                img = card.find('img')
                employee['image'] = urljoin(page_url, img['src']) if img and 'src' in img.attrs else None
                
                # استخراج الروابط الاجتماعية
                social_links = {}
                for a in card.find_all('a', href=True):
                    href = a['href']
                    if 'linkedin.com' in href:
                        social_links['linkedin'] = href
                    elif 'twitter.com' in href:
                        social_links['twitter'] = href
                    elif 'facebook.com' in href:
                        social_links['facebook'] = href
                
                employee['social_media'] = social_links
                
                # استخراج البريد الإلكتروني
                email = self.extract_email_from_card(card)
                employee['email'] = email
                
                if employee['name']:
                    self.results['employees'].append(employee)
                    
            except Exception as e:
                print(f"Error parsing employee card: {e}")
                continue

    def extract_email_from_card(self, card):
        # البحث عن البريد الإلكتروني في النص
        email_pattern = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
        text = card.get_text()
        emails = re.findall(email_pattern, text, re.IGNORECASE)
        
        if emails:
            return emails[0]
            
        # البحث عن روابط mailto:
        mailto = card.find('a', href=re.compile('mailto:'))
        if mailto and 'href' in mailto.attrs:
            return mailto['href'].replace('mailto:', '')
            
        return None

    def scrape_linkedin_profiles(self):
        if not self.driver:
            return
            
        print("[*] البحث عن ملفات LinkedIn للموظفين...")
        
        for employee in self.results['employees']:
            try:
                if not employee.get('social_media', {}).get('linkedin'):
                    # البحث في LinkedIn إذا لم يكن الرابط متوفرًا
                    search_query = f"{employee['name']} {self.results['company_info']['name']}"
                    search_url = f"https://www.linkedin.com/search/results/all/?keywords={search_query}"
                    
                    self.driver.get(search_url)
                    time.sleep(random.uniform(2, 4))
                    
                    # استخراج أول نتيجة بحث
                    first_result = self.driver.find_elements(By.CSS_SELECTOR, 'li.reusable-search__result-container')
                    if first_result:
                        profile_url = first_result[0].find_element(By.CSS_SELECTOR, 'a.app-aware-link').get_attribute('href')
                        employee['social_media']['linkedin'] = profile_url
                        
            except Exception as e:
                print(f"Error scraping LinkedIn for {employee['name']}: {e}")
                continue

    def find_employee_sites(self):
        print("[*] البحث عن مواقع الموظفين المرتبطين...")
        
        for employee in self.results['employees']:
            try:
                if not employee.get('personal_site'):
                    # البحث في حسابات وسائل التواصل الاجتماعي عن مواقع شخصية
                    if employee.get('social_media', {}).get('twitter'):
                        twitter_url = employee['social_media']['twitter']
                        twitter_bio = self.get_twitter_bio(twitter_url)
                        if twitter_bio:
                            employee['twitter_bio'] = twitter_bio
                            site = self.extract_url_from_text(twitter_bio)
                            if site:
                                employee['personal_site'] = site
                                
                    # إذا كان هناك حساب LinkedIn، ابحث في قسم "حول"
                    if not employee.get('personal_site') and employee.get('social_media', {}).get('linkedin'):
                        linkedin_url = employee['social_media']['linkedin']
                        linkedin_info = self.get_linkedin_info(linkedin_url)
                        if linkedin_info:
                            employee.update(linkedin_info)
                            
            except Exception as e:
                print(f"Error finding personal site for {employee['name']}: {e}")
                continue

    def get_twitter_bio(self, twitter_url):
        if not self.driver:
            return None
            
        try:
            self.driver.get(twitter_url)
            time.sleep(random.uniform(3, 5))
            
            bio = self.driver.find_element(By.CSS_SELECTOR, 'div[data-testid="UserDescription"]').text
            return bio
        except:
            return None

    def get_linkedin_info(self, linkedin_url):
        if not self.driver:
            return None
            
        try:
            self.driver.get(linkedin_url)
            time.sleep(random.uniform(4, 6))
            
            info = {}
            
            # الحصول على قسم "حول"
            try:
                about_section = self.driver.find_element(By.ID, 'about').find_element(By.XPATH, '..').text
                info['about'] = about_section
                
                # استخراج موقع الويب من قسم "حول"
                site = self.extract_url_from_text(about_section)
                if site:
                    info['personal_site'] = site
            except:
                pass
                
            return info
        except:
            return None

    def extract_url_from_text(self, text):
        url_pattern = r'https?://[^\s]+'
        urls = re.findall(url_pattern, text)
        
        if urls:
            # تصفية روابط وسائل التواصل الاجتماعي
            for url in urls:
                if not any(social in url for social in ['twitter.com', 'linkedin.com', 'facebook.com']):
                    return url
                    
        return None

    def analyze_technologies(self):
        print("[*] تحليل التقنيات المستخدمة...")
        try:
            response = requests.get(self.target_url, timeout=5)
            soup = BeautifulSoup(response.text, 'html.parser')
            
            # الكشف عن نظام إدارة المحتوى
            cms = self.detect_cms(soup)
            if cms:
                self.results['technologies'].append({
                    'type': 'CMS',
                    'name': cms,
                    'confidence': 'high'
                })
                
            # الكشف عن أطر العمل والتقنيات الأخرى
            tech_signatures = {
                'React': ['react.min.js', 'ReactDOM'],
                'Angular': ['angular.min.js', 'ng-app'],
                'Vue.js': ['vue.min.js', 'v-bind'],
                'jQuery': ['jquery.min.js', 'jQuery'],
                'Bootstrap': ['bootstrap.min.js', 'data-bs-toggle']
            }
            
            for tech, signatures in tech_signatures.items():
                for sig in signatures:
                    if sig in response.text:
                        self.results['technologies'].append({
                            'type': 'JavaScript Framework',
                            'name': tech,
                            'confidence': 'medium'
                        })
                        break
                        
        except Exception as e:
            print(f"Error analyzing technologies: {e}")

    def detect_cms(self, soup):
        cms_indicators = {
            'WordPress': ['wp-content', 'wp-includes', 'wordpress'],
            'Joomla': ['joomla', 'media/jui'],
            'Drupal': ['sites/all', 'drupal.js'],
            'Magento': ['magento/', 'skin/frontend'],
            'Shopify': ['cdn.shopify.com', 'shopify.js']
        }
        
        html = str(soup).lower()
        
        for cms, indicators in cms_indicators.items():
            if any(indicator.lower() in html for indicator in indicators):
                return cms
                
        return None

    def save_results(self, format='excel'):
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"employee_analysis_{self.domain}_{timestamp}"
        
        if format == 'excel':
            filename += '.xlsx'
            self.save_to_excel(filename)
        else:
            filename += '.json'
            with open(filename, 'w', encoding='utf-8') as f:
                json.dump(self.results, f, ensure_ascii=False, indent=4)
                
        print(f"[+] تم حفظ النتائج في ملف: {filename}")

    def save_to_excel(self, filename):
        wb = Workbook()
        
        # ورقة الموظفين
        ws_employees = wb.active
        ws_employees.title = "الموظفون"
        
        headers = ['الاسم', 'المنصب', 'البريد الإلكتروني', 'موقع شخصي', 'LinkedIn', 'Twitter', 'ملاحظات']
        ws_employees.append(headers)
        
        for emp in self.results['employees']:
            row = [
                emp.get('name', ''),
                emp.get('position', ''),
                emp.get('email', ''),
                emp.get('personal_site', ''),
                emp.get('social_media', {}).get('linkedin', ''),
                emp.get('social_media', {}).get('twitter', ''),
                emp.get('notes', '')
            ]
            ws_employees.append(row)
        
        # ورقة معلومات الشركة
        ws_company = wb.create_sheet("معلومات الشركة")
        ws_company.append(['اسم الشركة', self.results['company_info'].get('name', '')])
        ws_company.append(['الوصف', self.results['company_info'].get('description', '')])
        ws_company.append(['صفحة "من نحن"', self.results['company_info'].get('about_page', '')])
        
        # ورقة التقنيات
        ws_tech = wb.create_sheet("التقنيات")
        ws_tech.append(['النوع', 'الاسم', 'مستوى الثقة'])
        
        for tech in self.results['technologies']:
            ws_tech.append([
                tech.get('type', ''),
                tech.get('name', ''),
                tech.get('confidence', '')
            ])
        
        wb.save(filename)

    def run_analysis(self):
        print(f"[*] بدء تحليل الموقع: {self.target_url}")
        
        self.get_company_info()
        self.find_team_page()
        self.scrape_linkedin_profiles()
        self.find_employee_sites()
        self.analyze_technologies()
        
        if self.driver:
            self.driver.quit()
            
        print("[+] اكتمل التحليل بنجاح!")
        return self.results

if __name__ == "__main__":
    print("""
    أداة تحليل الموظفين المرتبطين بموقع إلكتروني
    -----------------------------------------
    هذه الأداة لأغراض بحثية وأمنية مشروعة فقط.
    """)
    
    if len(sys.argv) < 2:
        print("طريقة الاستخدام: python employee_analyzer.py <url>")
        sys.exit(1)
        
    target_url = sys.argv[1]
    
    analyzer = EmployeeSiteAnalyzer(target_url)
    results = analyzer.run_analysis()
    
    # عرض ملخص النتائج
    print("\nملخص النتائج:")
    print(f"- اسم الشركة: {results['company_info'].get('name', 'غير معروف')}")
    print(f"- عدد الموظفين المكتشفين: {len(results['employees'])}")
    
    if results['employees']:
        print("\nأبرز الموظفين:")
        for emp in results['employees'][:5]:  # عرض أول 5 موظفين فقط
            print(f"  - {emp.get('name')} ({emp.get('position')})")
            if emp.get('personal_site'):
                print(f"    موقع شخصي: {emp['personal_site']}")
    
    print("\nالتقنيات المستخدمة:")
    for tech in results['technologies']:
        print(f"  - {tech['name']} ({tech['type']})")
    
    # حفظ النتائج
    save_format = input("\nهل تريد حفظ النتائج بصيغة Excel (E) أو JSON (J)؟ ").lower()
    if save_format in ['e', 'excel']:
        analyzer.save_results(format='excel')
    elif save_format in ['j', 'json']:
        analyzer.save_results(format='json')
    else:
        print("لم يتم حفظ النتائج.")